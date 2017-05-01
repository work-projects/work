'''
Created on Apr 30, 2017

@author: Mukadam
'''
from selenium import webdriver
from openpyxl import load_workbook
import time
wb = load_workbook(filename = 'Paul_Xero_v10.xlsx')
sheet = wb.active
driver = webdriver.Firefox()
from selenium.webdriver.common.keys import Keys
try:
    driver.get('https://login.xero.com/')
    print 'On login page'
    driver.find_element_by_id('email').send_keys('tasneem.rangwala05@gmail.com')
    driver.find_element_by_id('password').send_keys('muidsa1!2@')
    driver.find_element_by_id('submitButton').click()
    print 'Clicked Submit '
    time.sleep(5)
    driver.get('https://go.xero.com/Accounts/Inventory')
    print ' In inventory'
    for i in range(2,3381):
        print sheet.cell(column=1, row=i).value
        while True:
            try:
                #print 'trying'
                driver.find_element_by_id('button-1012').click()
                break
            except:
                pass
        
        driver.find_element_by_name('Code').send_keys(sheet.cell(column=1, row=i).value)
        driver.find_element_by_name('Name').send_keys(sheet.cell(column=2, row=i).value)
        driver.find_element_by_name('PurchasesUnitPrice').send_keys(str(sheet.cell(column=4, row=i).value))
        driver.find_element_by_name('PurchasesAccountID').send_keys(sheet.cell(column=5, row=i).value)
        #time.sleep(2)
        driver.find_element_by_name('PurchasesAccountID').send_keys(Keys.RETURN)
#         while True:
#             try:
#                 driver.find_element_by_xpath(".//li[text()='Expenses']/../li[@class='x-boundlist-item']").click()
#                 break
#             except Exception,e:
#                 pass
        driver.find_element_by_name('PurchasesDescription').send_keys(sheet.cell(column=3, row=i).value)
        driver.find_element_by_name('SalesUnitPrice').send_keys(str(sheet.cell(column=8, row=i).value))
        driver.find_element_by_name('SalesAccountID').send_keys(sheet.cell(column=9, row=i).value)
        driver.find_element_by_name('SalesAccountID').send_keys(Keys.RETURN)
        #driver.find_element_by_xpath(".//li[text()='Revenue']/../li[@class='x-boundlist-item']").click()
        driver.find_element_by_name('SalesDescription').send_keys(sheet.cell(column=7, row=i).value)
        driver.find_element_by_name('SalesDescription').send_keys(Keys.TAB)
        driver.find_element_by_xpath('//*[text()="Save"]').send_keys(Keys.RETURN)
    
        while True:
            try:
                driver.find_element_by_id('inventoryitem-popup-body')
            except Exception, e:
                #print e
                break
    
    
    time.sleep(2)

except Exception,e:
    print e
finally:
    
    driver.quit()


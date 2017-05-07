'''
Created on May 5, 2017

@author: Mukadam
'''
from openpyxl import load_workbook
from shutil import copyfile
wb = load_workbook(filename = 'Input_before_men.xlsx')
sheets = wb.active
copyfile('Images_before_men_input/00a6af8319d8cb7410fffe1d299678860a1be8b35e73aeb4d788e7920c44fd31.jpeg', 'Images_before_men_output/abc.jpeg')
#print sheets.max_row
i=1
while i <= sheets.max_row:
    print i
    copyfile('Images_before_men_input/'+sheets.cell(column=2, row=i).value, 'Images_before_men_output/'+sheets.cell(column=1, row=i).value+'.jpeg')
    i+=1